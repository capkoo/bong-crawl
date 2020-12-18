import os
import sys
import datetime
import time

import openpyxl
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from worker.ImageDownloadWorker import ImageDowloadWorker


class BongImageDownloaderUI(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()


    def initUI(self):
        self.grid = QGridLayout()

        self.openFileBtn = QPushButton('&Open Excel')
        self.lineFilePath = QLineEdit('')
        self.actionButton = QPushButton('&Start')
        self.progressbar = QProgressBar()
        self.textResult = QTextEdit()
        self.tableWidget = QTableWidget()
        self.labelDelay = QLabel('딜레이')
        self.spinDelay = QDoubleSpinBox()

        self.lineFilePath.setReadOnly(True)
        self.actionButton.setDisabled(True)
        self.actionButton.clicked.connect(self.startDownload)
        self.openFileBtn.clicked.connect(self.showFileDialog)
        self.spinDelay.valueChanged.connect(self.delayChanged)
        self.progressbar.setValue(0)
        self.spinDelay.setValue(1.00)

        self.grid.addWidget(self.lineFilePath, 0, 0, 1, 4)
        self.grid.addWidget(self.openFileBtn, 0, 4, 1, 1)
        self.grid.addWidget(self.actionButton, 0, 5, 1, 1)
        self.grid.addWidget(self.labelDelay, 1, 0, 1, 1)
        self.grid.addWidget(self.spinDelay, 1, 1, 1, 1)
        self.grid.addWidget(self.progressbar, 2, 0, 1, 6)
        self.grid.addWidget(self.tableWidget, 3, 0, 1, 6)

        self.setLayout(self.grid)

        self.setWindowTitle('Bong Image Downloader')
        self.setGeometry(300, 100, 700, 600)
        self.show()


    def showFileDialog(self):
        fileDialog = QFileDialog()
        fname = fileDialog.getOpenFileName(self, 'Open file', './' ,"Excel Files (*.xls *.xlsx)")
        self.excelFilePath = fname[0]
        self.lineFilePath.setText(self.excelFilePath)

        if self.excelFilePath != '':
            self.actionButton.setDisabled(False)
        else:
            self.actionButton.setDisabled(True)
        try:
            self.setTableItem()
        except Exception as e:
            print(e)

    def delayChanged(self):
        self.delay = self.spinDelay.value()


    def setTableItem(self):
        self.tableWidget.clear()
        self.tableWidget.setRowCount(0)
        self.tableWidget.setColumnCount(0)
        self.progressbar.setValue(0)

        if self.excelFilePath == '':
            return

        wb = openpyxl.load_workbook(self.excelFilePath, read_only=True)
        ws = wb.active

        headerCells = ws['A1':'N1']
        headers = []
        for row in headerCells:
            for cell in row:
                headers.append(cell.value)

        self.tableWidget.setColumnCount(len(headers))
        self.tableWidget.setHorizontalHeaderLabels(headers)
        data = ws.iter_rows(min_row=2)

        rowCount = 0
        for x, rows in enumerate(data):
            if rows[0].value is not None:
                self.tableWidget.setRowCount(self.tableWidget.rowCount() + 1)
                rowCount += 1
                for y, cell in enumerate(rows):
                    val = cell.value
                    if val is not None:
                        item = QTableWidgetItem(str(val))
                        item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                        self.tableWidget.setItem(self.tableWidget.rowCount() - 1, y, item)

        self.progressbar.setMaximum(rowCount)



    def startDownload(self):
        self.progressbar.setValue(0)
        self.spinDelay.setDisabled(True)
        self.actionButton.setDisabled(True)
        self.actionButton.setText('&Cancel')
        self.actionButton.clicked.connect(self.stopDownload)
        self.actionButton.clicked.disconnect(self.startDownload)

        self.actionButton.setDisabled(False)

        self.openFileBtn.setDisabled(True)
        self.BASE_DIR = os.path.abspath('.');
        self.IMAGE_PATH = self.BASE_DIR + '/image/' + datetime.datetime.now().strftime('%Y%m%d')
        wb = openpyxl.load_workbook(self.excelFilePath, read_only=True)
        ws = wb.active
        data = ws.iter_rows(min_row=2)
        self.downloadWorker = ImageDowloadWorker()
        self.downloadWorker.setData(data)
        self.downloadWorker.setUi(self)
        self.downloadWorker.isRun = True
        self.downloadWorker.start()


    def stopDownload(self):
        try:
            self.downloadWorker.stop()
            QMessageBox.question(self, 'Message', '이미지 다운로드 취소!', QMessageBox.Yes)
            self.initWidget()
        except Exception as e:
            print(e)



    def initWidget(self):
        self.spinDelay.setDisabled(False)
        self.actionButton.setText('&Start')

        self.actionButton.clicked.disconnect(self.stopDownload)
        self.actionButton.clicked.connect(self.startDownload)
        self.actionButton.setDisabled(False)
        self.openFileBtn.setDisabled(False)


    def eventDoneDownload(self):
        QMessageBox.question(self, 'Message', '이미지 다운로드 완료!',QMessageBox.Yes)
        self.initWidget()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = BongImageDownloaderUI()
    sys.exit(app.exec_())